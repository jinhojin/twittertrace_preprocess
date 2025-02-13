import json
import re
import os
from openpyxl import Workbook

def parse_log_for_counts(log_file_path):

    ops_completed_pattern = re.compile(r"[0-9]*\.[0-9]*M ops completed")
    fifo_pattern = re.compile(r"FIFO rotation count becomes from (\d+) to (\d+)")

    counts = []
    current_count = 0

    if not os.path.isfile(log_file_path):
        return counts

    with open(log_file_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            line = line.strip()

            if ops_completed_pattern.search(line):
                current_count += 1

            if fifo_pattern.search(line):
                counts.append(current_count)
                current_count = 0


    return counts


def process_tasks_from_json(json_file_path, logs_dir, output_excel_file):
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    tasks_info = data.get("tasks", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "LSM_FifoRotation_Counts"

    ws.cell(row=1, column=1, value="Task")

    current_excel_row = 2

    for task_id, task_data in tasks_info.items():
        status_obj = task_data.get("status", {})
        if not isinstance(status_obj, dict):
            continue

        done_info = status_obj.get("Done")
        if not done_info:
            continue

        result = done_info.get("result", "")
        if result != "Success":
            continue

        label = task_data.get("label", "")
        if not label:
            continue
        if "lsm" not in label.lower():
            continue

        log_file_name = f"{task_id}.log"  
        log_file_path = os.path.join(logs_dir, log_file_name)

        counts = parse_log_for_counts(log_file_path)

        ws.cell(row=current_excel_row, column=1, value=label)

        for idx, cnt in enumerate(counts, start=2):
            ws.cell(row=current_excel_row, column=idx, value=cnt)

        current_excel_row += 1

    wb.save(output_excel_file)
    print(f"Excel: {output_excel_file}")


if __name__ == "__main__":

    json_path = "/root/.local/share/pueue/state.json"  
    logs_directory = "/root/.local/share/pueue/task_logs"
    output_excel_path = "/root/lsm_fifo_counts.xlsx"

    process_tasks_from_json(json_path, logs_directory, output_excel_path)
