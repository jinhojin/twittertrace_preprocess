import re
import sys
import os
import openpyxl
from openpyxl.styles import Font, Alignment

def parse_analysis_txt(file_path):
    """
    분석 결과 텍스트 파일을 읽어서,
    각 구역(Under 2KB, Over 2KB, All)에 대한 통계를 딕셔너리로 반환합니다.
    """
    data = {
        "Under2KB": {},
        "Over2KB": {},
        "All": {}
    }

    current_section = None

    patterns = {
        "Average key size"     : r"Average key size\s*:\s*([\d\.]+)",
        "Average value size"   : r"Average value size\s*:\s*([\d\.]+)",
        "Average object size"  : r"Average object size\s*:\s*([\d\.]+)",
        "Footprint1"           : r"Footprint1 \(sum of object size\)\s*:\s*([\d]+)",
        "Footprint2"           : r"Footprint2 \(sum of average of duplicated key\)\s*:\s*([\d]+)",
        "Unique key count"     : r"Unique key count\s*:\s*([\d]+)",
        "Total key count"      : r"Total key count\s*:\s*([\d]+)",
    }

    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        line = line.strip()
        if line.startswith("=== Under 2KB ==="):
            current_section = "Under2KB"
            continue
        elif line.startswith("=== Over 2KB ==="):
            current_section = "Over2KB"
            continue
        elif line.startswith("=== All ==="):
            current_section = "All"
            continue

        if not current_section:
            continue

        for key, pattern in patterns.items():
            m = re.search(pattern, line)
            if m:
                data[current_section][key] = m.group(1)

    return data

def convert_to_excel(data, output_xlsx):
    """
    질문에서 보여준 엑셀 레이아웃에 맞춰,
    data 딕셔너리를 엑셀로 변환하여 저장합니다.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Twitter trace_0"

    # 타이틀 행 병합
    ws.merge_cells("B1:D1")
    ws["B1"] = "Twitter trace_0"
    ws["B1"].font = Font(bold=True, size=14)
    ws["B1"].alignment = Alignment(horizontal="center")

    # 헤더
    ws.append(["", "<= 2KB", "> 2KB", "All"])

    def bytes_to_gb_str(val):
        return f"{int(val)/(1024**3):.2f}"

    def calc_percent(part, whole):
        return f"{part/whole*100:.2f}%" if whole != 0 else "0.00%"

    def with_commas(val):
        return f"{int(val):,}"

    def get(section, key, default="0"):
        return data[section].get(key, default)

    under2kb_fp1 = int(get("Under2KB","Footprint1","0"))
    over2kb_fp1  = int(get("Over2KB","Footprint1","0"))
    all_fp1      = int(get("All","Footprint1","0"))

    under2kb_fp1_gb = bytes_to_gb_str(under2kb_fp1)
    over2kb_fp1_gb  = bytes_to_gb_str(over2kb_fp1)
    all_fp1_gb      = bytes_to_gb_str(all_fp1)

    under2kb_fp1_pct = calc_percent(under2kb_fp1, all_fp1)
    over2kb_fp1_pct  = calc_percent(over2kb_fp1, all_fp1)

    under2kb_fp2 = int(get("Under2KB","Footprint2","0"))
    over2kb_fp2  = int(get("Over2KB","Footprint2","0"))
    all_fp2      = int(get("All","Footprint2","0"))

    under2kb_fp2_gb = bytes_to_gb_str(under2kb_fp2)
    over2kb_fp2_gb  = bytes_to_gb_str(over2kb_fp2)
    all_fp2_gb      = bytes_to_gb_str(all_fp2)

    under2kb_fp2_pct = calc_percent(under2kb_fp2, all_fp2)
    over2kb_fp2_pct  = calc_percent(over2kb_fp2, all_fp2)

    table_rows = [
        ["Average key size",
         get("Under2KB","Average key size","0"),
         get("Over2KB","Average key size","0"),
         get("All","Average key size","0")],
        ["Average value size",
         get("Under2KB","Average value size","0"),
         get("Over2KB","Average value size","0"),
         get("All","Average value size","0")],
        ["Average object size",
         get("Under2KB","Average object size","0"),
         get("Over2KB","Average object size","0"),
         get("All","Average object size","0")],

        ["Sum of the size of all objects (GB)",
         under2kb_fp1_gb, over2kb_fp1_gb, all_fp1_gb],
        ["",
         under2kb_fp1_pct, over2kb_fp1_pct, "100%"],

        ["Sum of the average size of objects (GB)",
         under2kb_fp2_gb, over2kb_fp2_gb, all_fp2_gb],
        ["",
         under2kb_fp2_pct, over2kb_fp2_pct, "100%"],

        ["# of unique keys",
         with_commas(get("Under2KB","Unique key count","0")),
         with_commas(get("Over2KB","Unique key count","0")),
         with_commas(get("All","Unique key count","0"))],

        ["# of total objects",
         with_commas(get("Under2KB","Total key count","0")),
         with_commas(get("Over2KB","Total key count","0")),
         with_commas(get("All","Total key count","0"))]
    ]

    for row_data in table_rows:
        ws.append(row_data)

    ws.column_dimensions["A"].width = 45
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 16

    wb.save(output_xlsx)
    print(f"✅ 엑셀 파일이 생성되었습니다: {output_xlsx}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python script.py [input_txt]")
        sys.exit(1)

    input_txt = sys.argv[1]
    # 확장자를 제외한 파일명 + ".xlsx" 만들기
    output_xlsx = os.path.splitext(input_txt)[0] + ".xlsx"

    # 1) 텍스트 파일 파싱
    parsed_data = parse_analysis_txt(input_txt)
    # 2) 엑셀 생성
    convert_to_excel(parsed_data, output_xlsx)
